import os, re, sys
from dotenv import load_dotenv
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
#Selenium imports
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import traceback
import pandas as pd
import time
from openpyxl import Workbook, load_workbook
from selenium.webdriver.chrome.options import Options
from screeninfo import get_monitors
from datetime import datetime
from selenium.common.exceptions import NoSuchElementException


load_dotenv()

class ConsultTrainees:

    def __init__(self, user=None, password=None) -> None:

         #Carpeta raíz
        self.current_root_path = os.getcwd()
        
        #Environ Variables
        self.default_url = "http://senasofiaplus.edu.co/sofia-public/"
        self.default_user = os.environ.get("DEFAULT_USER")
        self.default_pass = os.environ.get("DEFAULT_PASS")

        #Initialize webdriver
        self.options = Options()
        prefs = {"download.default_directory" : f"{self.current_root_path}/procesar/"}
        self.options.add_experimental_option("prefs", prefs)
        self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=self.options)
        self.wait = WebDriverWait(self.driver, 50)

       
        
        principal_folders = ["/entrada/", '/procesar/', '/reporte/']
        for folder in principal_folders: # Se crean las carpetas necesarias
            if not os.path.exists(f"{self.current_root_path}{folder}"):
                os.mkdir(f"{self.current_root_path}{folder}")
        self.path_process_folder = f"{self.current_root_path}/procesar/"
        
        self.df_consolidated = pd.DataFrame()



    def open_webdriver(self, url=None):
        try:
            if not url:
                url = self.default_url
            self.driver.get(url)
           
            
            # get screen size
            monitor = get_monitors()[0]
            screen_width = monitor.width
            screen_height = monitor.height

            self.driver.set_window_size(screen_width, 550)

            # calculate position for top right corner
            x_position = 0  # left edge
            y_position = screen_height - self.driver.get_window_size()['height']

            # set window position
            self.driver.set_window_position(x_position, y_position)
            self.driver.switch_to.frame(0)
        except Exception as err:
            print(err)
            print(traceback.format_exc())
    

    def login_process(self, user=None, password=None):
        try:
            if not user and not password:
                user =  self.default_user
                password = self.default_pass
            if user == None or password == None:
                print('ERROR EN PROCESO DE LOGIN - POR FAVOR DIGITE CREDENCIALES VALIDAS')
                return False

            try:
                element = self.wait.until(EC.element_to_be_clickable((By.ID, "username")))
                element.click()
                element.send_keys(user)
                element = self.wait.until(EC.element_to_be_clickable((By.NAME, "josso_password")))
                element.click()
                element.send_keys(password)
                element = self.wait.until(EC.element_to_be_clickable((By.NAME, "ingresar")))
                element.click()
                print('LOGGEO EXITOSO!')
                return True
            except Exception as err:
                print('ERROR EN PROCESO DE LOGIN - POR FAVOR DIGITE CREDENCIALES VALIDAS')
                #print(traceback.format_exc())
                return False

        except Exception as err:
            print('ERROR EN PROCESO DE LOGIN - POR FAVOR DIGITE CREDENCIALES VALIDAS')
            return False
            # print(err)
            # print(traceback.format_exc())


    def select_role(self):
        try:
            self.driver.execute_script("return self.name || self.id")
            element = self.driver.find_element(By.XPATH, '//*[@id="seleccionRol:roles"]') 
            self.driver.execute_script("arguments[0].click();", element)
            element = self.wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="seleccionRol:roles"]/option[6]')))
            element.click()
            print("ROL SELECCIONADO")
        except Exception as err:
            try:
                element = self.driver.find_element(By.XPATH, '//*[@id="j_id_jsp_1108549618_6:dtprogramacionesDeAmbiente:0:j_id_jsp_1108549618_10"]')
                self.driver.execute_script("arguments[0].click();", element)
            except Exception as err:
                print("ERROR AL SELECCIONAR ROL")
                # print(err)
                # print(traceback.format_exc())           


    def obtain_fichas_a_descargar(self, 
                               document_name=None, 
                               other_path=None,  #other path se va
                               extraPath_toRoot="/entrada/"):
        
        #Cargar archivo de consulta
        #Verificar si ya existe un archivo de consulta, seleccionar si utilizar ese o reemplazarlo

        """
        Descarga de fichas...

        Attributes
            document_name: nombre archivo excel donde las fichas
            other_path (opcional): En caso de que el archivo no se encuentre en root 
            extraPath_toRoot (default): en caso de que se encuentre dentro de otras carpetas

        """
        
        try:
            # Check if the directory is empty
            if not os.listdir(self.current_root_path + extraPath_toRoot):
                print("NO SE HA SUBIDO UN ARCHIVO DE ENTRADA - POR FAVOR REALIZAR LA CARGA")
                return False
            #Check if there is a document on init folder
            if document_name is None and os.listdir(self.current_root_path + extraPath_toRoot):
                document_name = os.listdir(self.current_root_path + extraPath_toRoot)[0]

            consult_path  = f"{self.current_root_path}{extraPath_toRoot}{document_name}"
            if other_path:
                consult_path  = f"{other_path}{document_name}"

            #Get all fichas
            if os.path.exists(consult_path):
                dfBase = pd.read_excel(consult_path)
                if 'FICHAS' not in dfBase.columns:
                    print('ERROR AL LEER EL ARCHIVO A PROCESAR, POR FAVOR CARGAR UN ARCHIVO VALIDO')
                    return False
                dfBase = pd.DataFrame(dfBase['FICHAS'])
                dfBase.drop_duplicates(subset=['FICHAS'], inplace=True)
                identificador_ficha_list = dfBase['FICHAS'].tolist()
                print("FICHAS INGRESADAS PARA PROCESAR", identificador_ficha_list)
                return identificador_ficha_list

        except Exception as err:
            print('ERROR EN PROCESO DE OBTENCIÓN DE FICHAS A DESCARGAR')
            print(err)
            print(traceback.format_exc())
            return False

    def keep_driver_open(self):
        input("Press any key to close the webdriver...")
        self.driver.quit()


    def depurate_from_existing_files(self, 
                               extraPath_toRoot="/procesar/",
                               list_fichas=None):
        

        """
        Calcular la cantidad de archivos faltantes y mostrarla

        Attributes
            other_path (opcional): En caso de que el archivo no se encuentre en root
            extraPath_toRoot (default): En caso de que se encuentre dentro de otras carpetas
            list_fichas []: Lista de las fichas obtenidas de entrada
        """
        try:
            removed_fichas = []
            procesar_path  = f"{self.current_root_path}{extraPath_toRoot}"
            file_list = os.listdir(procesar_path)
            existing_files = [file for file in file_list if file.startswith('Reporte de Aprendices Ficha')]
            if len(existing_files) == 0:
                existing_files = [file for file in file_list if file.startswith('Reporte de Juicios Evaluativos')]
            # Calcular la cantidad de archivos faltantes y mostrarla
            for f in existing_files:
                existing_ficha = re.findall(r'\d+', f)
                if int(existing_ficha[0]) in list_fichas:
                    list_fichas.pop(list_fichas.index(int(existing_ficha[0])))
                    removed_fichas.append(int(existing_ficha[0]))
            #print(f"Hay {len(list_fichas)} archivos faltantes por descargar")
            print("Fichas a descargar: ", list_fichas)
            print("Fichas encontradas que ya han sido descargadas: ", removed_fichas)
            return list_fichas
        except Exception as err:
            print('ERROR EN EL PROCESO DE DEPURACIÓN FICHAS')
            return False
            # print(err)
            # print(traceback.format_exc())


    def download_files(self, list_fichas):
        try:
            if len(list_fichas) == 0:
                print("No hay fichas por descargar...")
                return False
            for f in list_fichas:
                self.driver.switch_to.default_content()
                element = self.wait.until(EC.element_to_be_clickable((By.XPATH, f'//body/div/div/nav/div[2]/div/div/form[2]/ul/li[6]/a')))
                #element.click()
                element = self.driver.find_element(By.XPATH, f'//body/div/div/nav/div[2]/div/div/form[2]/ul/li[6]/a')
                self.driver.execute_script("arguments[0].click();", element)
                element = self.driver.find_element(By.XPATH, f'//body/div/div/nav/div[2]/div/div/form[2]/ul/li[6]/ul/li/a')
                self.driver.execute_script("arguments[0].click();", element)
                element = self.driver.find_element(By.XPATH, f'//body/div/div/nav/div[2]/div/div/form[2]/ul/li[6]/ul/li/ul/li[3]/ul/li[3]/a')
                self.driver.execute_script("arguments[0].click();", element)
                self.driver.switch_to.frame(0)

                # Buscar la ficha de caracterización correspondiente y descargar el informe
                element = self.driver.find_element(By.XPATH, f'//body/div[2]/form/fieldset/div/table/tbody/tr/td[2]/table/tbody/tr/td[2]/a')
                time.sleep(1)
                element.click()
                self.driver.switch_to.frame(0)
                element = self.driver.find_element(By.XPATH, f'//div[2]/form/fieldset/div/table/tbody/tr/td[2]/input')
                element.send_keys(f)
                #print(len(nuevo_df_sin_duplicados), "por descargar")
                element = self.driver.find_element(By.XPATH, f'//div[2]/form/fieldset/div/div/input')
                self.driver.execute_script("arguments[0].click();", element)

                #check if ficha its correct
                time.sleep(1)
                try:
                    elemento2 = self.driver.find_element(By.XPATH,f'/html/body[1]/div[2]/table/tbody/tr/td/span')
                except NoSuchElementException:
                    print(f'No se encontraron registros para la ficha ({f}) seleccionada... Por favor verificar')
                    elemento2 = None
                if elemento2:
                    print(f'No se encontraron registros para la ficha ({f}) seleccionada... Por favor verificar')
                    return False
                else:
                    element = self.driver.find_element(By.XPATH, f'//body/div[2]/form/div/fieldset/table/tbody/tr/td/a')
                    self.driver.execute_script("arguments[0].click();", element)
                    self.driver.switch_to.default_content()
                    self.driver.switch_to.frame(0)
                    element = self.driver.find_element(By.XPATH, f'//body/div[2]/form/fieldset/div/div/input')
                    self.driver.execute_script("arguments[0].click();", element)
                    while os.path.exists(f"{self.current_root_path}/procesar/Reporte de Aprendices Ficha {f}.xls.crdownload") or \
                        any(file.startswith(".com.google.Chrome") for file in os.listdir(self.current_root_path)):
                        print("Esperando a que el archivo termine de descargarse...")
                        time.sleep(1)
                    if os.path.exists(f"{self.current_root_path}/procesar/Reporte de Aprendices Ficha {f}.xls"):
                        print(f"Descarga finalizada de {f}")
        except Exception as err:
            print('ERROR EN LA DESCARGA FICHAS - INTENTE NUEVAMENTE')
            # print(err)
            # print(traceback.format_exc())
            return False
            


    def modified_files(self):
        try:
            for filename in os.listdir(self.path_process_folder):
                if filename.endswith('.xls'):
                    print('Procesando...', filename)
                    name, extension = os.path.splitext(filename)
                    file_path_xlsx = f"{self.current_root_path}/procesar/{name}.xlsx"
                    df = pd.read_excel(f"{self.current_root_path}/procesar/{name}{extension}")
                    df.to_excel(file_path_xlsx, index=False)
                    os.remove(f"{self.current_root_path}/procesar/{filename}")
                    # Load the workbook
                    book = load_workbook(file_path_xlsx)
                    sheet = book.active

                    # Get the cell values
                    a2 = sheet.cell(row=2, column=1).value
                    c2 = sheet.cell(row=2, column=3).value
                    a3 = sheet.cell(row=3, column=1).value
                    c3 = sheet.cell(row=3, column=3).value

                    # Create a new workbook and sheet
                    new_book = Workbook()
                    new_sheet = new_book.active

                    # Write the values of the first 5 rows to the new sheet
                    for row in range(1, 6):
                        for col in range(1, sheet.max_column + 1):
                            new_sheet.cell(row=row, column=col+2).value = sheet.cell(row=row, column=col).value

                    # Write the header in the new column
                    new_sheet.cell(row=5, column=1).value = a2
                    new_sheet.cell(row=5, column=2).value = a3

                    # Write the values for the rest of the columns
                    for row in range(6, sheet.max_row + 1):
                        new_sheet.cell(row=row, column=1).value = c2
                        new_sheet.cell(row=row, column=2).value = c3
                        for col in range(1, sheet.max_column + 1):
                            new_sheet.cell(row=row, column=col+2).value = sheet.cell(row=row, column=col).value

                    # Save the modified Excel file
                    file_path_xlsx
                    new_file_path = os.path.join(self.path_process_folder, 'modified_' + name + '.xlsx')
                    new_book.save(new_file_path)

                    # Read the modified file and add the contents to the DataFrame
                    df_file = pd.read_excel(new_file_path, index_col=0)
                    self.df_consolidated = pd.concat([self.df_consolidated, df_file], ignore_index=True)

                    df_file_without_header = pd.read_excel(new_file_path)
                    df_file_without_header.to_excel(new_file_path, index=False)

        except Exception as err:
            print('ERROR EN AL INTENTAR PROCESAR LOS ARCHIVOS - POR FAVOR INTENTE DE NUEVO')
            # print(err)
            # print(traceback.format_exc())

    def generate_consolidated_trainees(self):
        try: 
            for filename in os.listdir(self.path_process_folder):
                if filename.startswith('modified') and (filename.endswith('.xls') or filename.endswith('.xlsx')):
                    file_path = os.path.join(self.path_process_folder, filename)

                    # Leer el archivo de Excel
                    df = pd.read_excel(file_path, header=4)

                    self.df_consolidated = pd.concat([self.df_consolidated, df], ignore_index=True)
                    
            # Filtrar las filas que contengan 'formación' (en mayúsculas o minúsculas)
            df_formacion = self.df_consolidated[self.df_consolidated.apply(lambda row: 'formación' in str(row).lower(), axis=1)]

            # Guardar el DataFrame en un archivo de Excel
            output_file_path = self.path_process_folder + '/consolidado_final.xlsx'
            df_formacion.to_excel(output_file_path, index=False)

            # print(f"Se ha guardado el archivo consolidado_final.xlsx en la siguiente ruta: {output_file_path}")
            file_list = [f for f in os.listdir(self.path_process_folder) if f.startswith('modified')]

            df_list = []
            for file_name in file_list:
                file_path = os.path.join(self.path_process_folder, file_name)
                print("Procesando archivo: ", file_name)
                df = pd.read_excel(file_path, skiprows=4)
                df = df[['Ficha de Caracterización:', 'Estado:', 'Tipo de Documento', 'Número de Documento', 'Nombre', 'Apellidos', 'Celular', 'Correo Electrónico', 'Estado']]
                df_list.append(df)

            df_concat = pd.concat(df_list, axis=0, ignore_index=True)

            today = datetime.today()
            date_string = today.strftime('%Y-%m-%d--%H:%M')
        
            output_path = f"{self.current_root_path}/reporte/Consolidado_Reporte_Aprendices_{date_string}.xlsx"
            df_concat.to_excel(output_path, index=False)
            df_concat = pd.read_excel(output_path)

            # Ver los primeros 5 registros del DataFrame
            #print(df_concat.head())
            df_filtered = df_concat[df_concat['Estado'] == 'EN FORMACIÓN']
        
            # Generar un archivo de Excel en la ruta deseada
            print("GENERACIÓN DE CONSOLIDADO EXITOSA")
            sys.stdout.flush()
            output_path = self.path_process_folder + "/Datos_formacion.xlsx"
            df_filtered.to_excel(output_path, index=False)
            # Ver los primeros 5 registros del DataFrame filtrado
            #print(df_filtered.head())
            texto_df = pd.DataFrame(df_filtered.iloc[:, 0].str.split('-', expand=True)[0])

            return output_path
            
        except Exception as err:
            print('ERROR EN AL INTENTAR GENERAR ARCHIVO CONSOLIDADO - POR FAVOR INTENTE DE NUEVO')
            # print(err)
            # print(traceback.format_exc())
            return False
        

    def delete_all_files_in_procesar_and_entrada(self, confirmation):
        try:
            if confirmation:
                print(confirmation)
                for filename in os.listdir(f"{self.current_root_path}/procesar",):
                    file_path = os.path.join(f"{self.current_root_path}/procesar", filename)
                    try:
                        if os.path.isfile(file_path):
                            os.unlink(file_path)
                    except Exception as e:
                        print(f'Failed to delete {file_path}. Reason: {e}')
                for filename in os.listdir(f"{self.current_root_path}/entrada",):
                    file_path = os.path.join(f"{self.current_root_path}/entrada", filename)
                    try:
                        if os.path.isfile(file_path):
                            os.unlink(file_path)
                    except Exception as e:
                        print(f'Failed to delete {file_path}. Reason: {e}')
            else:
                print('HA OCURRIDO UN ERROR EN EL PROCESO... EJECUTE NUEVAMENTE')
        except Exception as err:
            print('HA OCURRIDO UN ERROR EN EL PROCESO... EJECUTE NUEVAMENTE')
            return False
            print(err)
            print(traceback.format_exc())


    def download_juicios_process(self, list_fichas):
        try:
            if len(list_fichas) == 0:
                print("No hay fichas por descargar...")
                return False
            for f in list_fichas:
                self.driver.switch_to.default_content()
                elemento = self.wait.until(EC.element_to_be_clickable((By.XPATH, f'//*[@id="side-menu"]/li[6]/a')))
                self.driver.execute_script("arguments[0].click();", elemento)
                elemento = self.driver.find_element(By.XPATH, f'//*[@id="side-menu"]/li[6]/ul/li[1]/a')
                self.driver.execute_script("arguments[0].click();", elemento)
                elemento = self.driver.find_element(By.XPATH, f'//*[@id="side-menu"]/li[6]/ul/li[1]/ul/li[3]/a')
                self.driver.execute_script("arguments[0].click();", elemento)
                elemento = self.driver.find_element(By.XPATH, f'//*[@id="146979Opcion"]')
                self.driver.execute_script("arguments[0].click();", elemento)
                self.driver.switch_to.frame('contenido')
                time.sleep(1)
                elemento = self.driver.find_element(By.XPATH, f'/html/body[1]/div[2]/form/fieldset/div/table/tbody/tr/td[2]/table/tbody/tr/td[2]/a')
                self.driver.execute_script("arguments[0].click();", elemento)
                self.driver.switch_to.frame('modalDialogContentviewDialog2')
                time.sleep(1)
                elemento = self.driver.find_element(By.XPATH, f'//*[@id="form:codigoFichaITX"]')
                self.driver.execute_script("arguments[0].click();", elemento)
                elemento.send_keys(f) #//*[@id="j_id_jsp_581860097_102"]
                elemento = self.driver.find_element(By.XPATH,f'//div[2]/form/fieldset/div/div/input') 
                self.driver.execute_script("arguments[0].click();", elemento)
                
                #check if element id wrong
                time.sleep(1)
                try:
                    elemento2 = self.driver.find_element(By.XPATH,f'/html/body[1]/div[2]/table/tbody/tr/td/span')
                except NoSuchElementException:
                    print(f'No se encontraron registros para la ficha ({f}) seleccionada... Por favor verificar')
                    elemento2 = None
                if elemento2:
                    print(f'No se encontraron registros para la ficha ({f}) seleccionada... Por favor verificar')
                    return False
                else:
                    elemento = self.driver.find_element(By.XPATH,f'//body/div[2]/form/div/fieldset/table/tbody/tr/td/a')
                    self.driver.execute_script("arguments[0].click();", elemento)
                    self.driver.switch_to.default_content()
                    self.driver.switch_to.frame(0) 
                    elemento = self.driver.find_element(By.XPATH,f'//body/div[2]/form/fieldset/div/div/input')
                    self.driver.execute_script("arguments[0].click();", elemento)
                    time.sleep(1)
                    while os.path.exists(f"{self.current_root_path}/procesar/Reporte de Juicios Evaluativos.xls.crdownload") or \
                        any(file.startswith(".com.google.Chrome") for file in os.listdir(self.current_root_path)):
                        print("Esperando a que el archivo termine de descargarse...")
                        time.sleep(1)
                    if os.path.exists(f"{self.current_root_path}/procesar/Reporte de Juicios Evaluativos.xls"):
                        print(f"Descarga finalizada de Reporte de Juicios {f}")
                        os.rename(f"{self.current_root_path}/procesar/Reporte de Juicios Evaluativos.xls", \
                                f"{self.current_root_path}/procesar/Reporte de Juicios Evaluativos {f}.xls")
        except Exception as err:
            print('ERROR EN LA DESCARGA FICHAS - INTENTE NUEVAMENTE')
            print(err)
            print(traceback.format_exc())
            return False
            


    def restructure_for_consolidated_file(self):
        try:
            for filename in os.listdir(f"{self.current_root_path}/procesar/"):
                if filename.startswith('Reporte de Juicios Evaluativos') and filename.endswith('.xls'):
                    if filename.endswith('.xls'):
                        print('Procesando...', filename)
                        name, extension = os.path.splitext(filename)
                        file_path_xlsx = f"{self.current_root_path}/procesar/{name}.xlsx"
                        df = pd.read_excel(f"{self.current_root_path}/procesar/{name}{extension}")
                        df.to_excel(file_path_xlsx, index=False)
                        os.remove(f"{self.current_root_path}/procesar/{filename}")
                        # Load the workbook
                        book = load_workbook(file_path_xlsx)
                        sheet = book.active
  

                        # Create a new Excel workbook and a new worksheet
                        new_book = Workbook()
                        new_sheet = new_book.active

                        # Write the values to the new worksheet
                        new_sheet['A1'] = sheet['A3'].value
                        new_sheet['B1'] = sheet['A4'].value
                        new_sheet['C1'] = sheet['A6'].value
                        new_sheet['D1'] = sheet['A7'].value
                        new_sheet['E1'] = sheet['A8'].value
                        new_sheet['F1'] = sheet['A9'].value
                        new_sheet['G1'] = sheet['A11'].value
                        new_sheet['H1'] = sheet['A12'].value

                        #C's
                        c3 = sheet['C3'].value
                        c4 = sheet['C4'].value
                        c5 = sheet['C5'].value
                        c6 = sheet['C6'].value
                        c7 = sheet['C7'].value
                        c8 = sheet['C8'].value
                        c9 = sheet['C9'].value
                        c10 = sheet['C10'].value
                        c11 = sheet['C11'].value

                        # Write the values to the new worksheet
                        for row in range(2, sheet.max_row + 1):
                            if sheet.max_column > 9 and sheet.cell(row=row, column=9).value == '':
                                new_sheet.cell(row=row, column=9, value=c8)
                            elif sheet.max_column > 10 and sheet.cell(row=row, column=10).value == '':
                                new_sheet.cell(row=row, column=10, value=c9)
                            else:
                                pass
                                # Haz algo si la celda ya tiene un valor
                            new_sheet.cell(row=row, column=1, value=c3)
                            new_sheet.cell(row=row, column=2, value=c4)
                            new_sheet.cell(row=row, column=3, value=c6)
                            new_sheet.cell(row=row, column=4, value=c7)
                            new_sheet.cell(row=row, column=5, value=c8)
                            new_sheet.cell(row=row, column=6, value=c9)
                            new_sheet.cell(row=row, column=7, value=c10)
                            new_sheet.cell(row=row, column=8, value=c11)


                        for row in range(13, sheet.max_row + 1):
                            for col in range(1, 10):
                                new_sheet.cell(row=row-12, column=col+8, value=sheet.cell(row=row, column=col).value)

                        # Get the original file name without the extension
                        filename_without_ext = os.path.splitext(filename)[0]

                        # Define the path of the new modified Excel file
                        new_file_path = os.path.join(f"{self.current_root_path}/procesar/", 'new_' + filename_without_ext + '.xlsx')
                        new_book.save(new_file_path)

        except Exception as err:
            print(err)
            print(traceback.format_exc())


    def generate_consolidated_juicios(self):
        try:
            today = datetime.today()
            date_string = today.strftime('%Y-%m-%d--%H:%M')

            # Create new Excel file
            new_book = Workbook()
            new_sheet = new_book.active
            new_sheet.title = 'Sheet 1'

            # Define column headers
            headers = ['Ficha de Caracterización', 'Cógigo', 'Denominación', 'Estado de la Ficha de Caracterización', 'Fecha Inicio', 'Fecha Fin', 'Regional', 'Centro de Formación', 'Tipo de Documento', 'Número de Documento', 'Nombre', 'Apellidos', 'Estado', 'Competencia', 'Resultado de Aprendizaje', 'Juicio de Evaluación', 'Funcionario que registro el juicio evaluativo']

            # Write headers in the new file
            for col in range(len(headers)):
                new_sheet.cell(row=1, column=col+1, value=headers[col])

            # Define path and name of the consolidated file
            consolidated_file_path = f'{self.current_root_path}/reporte/Consolidado_Reporte_Juicios_{date_string}.xlsx'

            # Combine all Excel files in the folder
            file_list = [f for f in os.listdir(self.path_process_folder) if f.startswith('new_')]
            for filename in file_list:
                # if filename.startswith('new_') and filename.endswith('.xlsx'):
                filepath = os.path.join(f'{self.current_root_path}/procesar/', filename)
                book = load_workbook(filepath)
                sheet = book.active

                # Write data in the new file
                for row in range(2, sheet.max_row + 1):
                    for col in range(1, sheet.max_column + 1):
                        new_sheet.cell(row=row, column=col, value=sheet.cell(row=row, column=col).value)

            # Save the consolidated file
            new_book.save(consolidated_file_path)
            print("SE HA CREADO EL CONSOLIDADO EXITOSAMENTE")
            return consolidated_file_path
        except Exception as err:
            return False
        
    

    

# BotConsulta = ConsultTrainees()
# BotConsulta.open_webdriver()
# BotConsulta.login_process()
# BotConsulta.select_role()
# list_fichas = BotConsulta.obtain_fichas_a_descargar()
# # # print(list_fichas)
# final_download_files = BotConsulta.depurate_from_existing_files(list_fichas=list_fichas)

# BotConsulta.download_juicios_process(final_download_files)
# BotConsulta.restructure_for_consolidated_file()
# BotConsulta.generate_consolidated_juicios()
# # BotConsulta.modified_files()
# # BotConsulta.generate_consolidated_trainees()
# # BotConsulta.keep_driver_open()

